import openpyxl

wb = openpyxl.load_workbook(r'D:\DEV\ExamTrack_Pro\Practical Schedule-2025-26 - Term-1(Even & Even Junior).xlsx', read_only=True, data_only=True)

with open(r'D:\DEV\ExamTrack_Pro\scripts\xlsx_output.txt', 'w', encoding='utf-8') as f:
    f.write(f'Sheets: {wb.sheetnames}\n\n')
    ws = wb['Practical']
    
    # Print first 10 rows to understand structure
    f.write("=== First 10 rows ===\n")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 10:
            f.write(f'Row {i+1}: {list(row)}\n')
        else:
            break
    
    # Now get unique subject code -> subject name mapping
    f.write("\n=== Unique Subject Code -> Subject Name mappings ===\n")
    subject_map = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c is not None else '' for c in row]
            f.write(f'Headers: {headers}\n')
            # Find column indices
            try:
                code_idx = headers.index('Subject Code')
                name_idx = headers.index('Subject Name')
                f.write(f'Subject Code col: {code_idx}, Subject Name col: {name_idx}\n\n')
            except ValueError as e:
                f.write(f'Error: {e}\nLooking for similar headers...\n')
                for j, h in enumerate(headers):
                    f.write(f'  col {j}: {h}\n')
            continue
        
        if i > 0 and headers:
            try:
                code = row[code_idx] if code_idx < len(row) else None
                name = row[name_idx] if name_idx < len(row) else None
                if code and name:
                    subject_map[str(name).strip()] = str(code).strip()
            except:
                pass
    
    for name, code in sorted(subject_map.items()):
        f.write(f'  "{name}": "{code}"\n')

print("Done! Check scripts/xlsx_output.txt")
