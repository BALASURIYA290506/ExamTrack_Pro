import openpyxl

XLSX_PATH = r'D:\DEV\Projects\ExamTrack_Pro\End Semester Theory & Practical Schedule(2025-26 Even & Even Junior -Term-2) - Final - To be Shared with the students.xlsx'

def main():
    try:
        wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
        print("Sheet names:", wb.sheetnames)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            row_count = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 5:  # print first 5 rows
                    print(f"Row {i}: {[str(c) for c in row]}")
                row_count += 1
            print(f"Total rows: {row_count}")
    except Exception as e:
        print("Error:", e)
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
