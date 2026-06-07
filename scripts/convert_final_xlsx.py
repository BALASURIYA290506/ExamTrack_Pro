import json
import openpyxl

XLSX_PATH = r'D:\DEV\Projects\ExamTrack_Pro\End Semester Theory & Practical Schedule(2025-26 Even & Even Junior -Term-2) - Final - To be Shared with the students.xlsx'
PRACTICAL_JSON = r'D:\DEV\Projects\ExamTrack_Pro\src\data\students_practical.json'
THEORY_JSON = r'D:\DEV\Projects\ExamTrack_Pro\src\data\students_theory.json'


def process_theory(ws):
    records = []
    headers = None
    skipped = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c is not None else '' for c in row]
            print(f"  Theory headers: {headers}")
            continue

        if all(cell is None for cell in row):
            skipped += 1
            continue

        row_data = dict(zip(headers, row))

        reg_no = row_data.get('Register Number')
        name = row_data.get('Name')
        date = row_data.get('Date')
        slot = row_data.get('FN/AN')
        code = row_data.get('Subject Code')
        subject = row_data.get('Subject Name')
        room = row_data.get('Room/Hall', '')

        if not reg_no or not name or not date or not subject:
            skipped += 1
            continue

        reg_str = str(int(reg_no)) if isinstance(reg_no, float) else str(reg_no).strip()

        record = {
            "Student Name":    str(name).strip(),
            "Register Number": reg_str,
            "Date":            str(date).strip() if date else "",
            "Slot":            str(slot).strip() if slot else "",
            "Category":        "Theory",
            "Subject Name":    str(subject).strip() if subject else "",
            "Room / Hall":     str(int(room)) if isinstance(room, (int, float)) and room else (str(room).strip() if room else ""),
            "Subject Code":    str(code).strip() if code else "",
        }
        records.append(record)

    return records, skipped


def process_practical(ws):
    records = []
    headers = None
    skipped = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c is not None else '' for c in row]
            print(f"  Practical headers: {headers}")
            continue

        if all(cell is None for cell in row):
            skipped += 1
            continue

        row_data = dict(zip(headers, row))

        reg_no = row_data.get('Register Number')
        name = row_data.get('Student Name')
        date = row_data.get('Date')
        slot = row_data.get('FN/AN')
        code = row_data.get('Course Code')
        subject = row_data.get('Course Name')
        location = row_data.get('Location', '')

        if not reg_no or not name or not date or not subject:
            skipped += 1
            continue

        reg_str = str(int(reg_no)) if isinstance(reg_no, float) else str(reg_no).strip()

        record = {
            "Student Name":    str(name).strip(),
            "Register Number": reg_str,
            "Date":            str(date).strip() if date else "",
            "Slot":            str(slot).strip() if slot else "",
            "Category":        "Practical",
            "Subject Name":    str(subject).strip() if subject else "",
            "Room / Hall":     str(int(location)) if isinstance(location, (int, float)) and location else (str(location).strip() if location else ""),
            "Subject Code":    str(code).strip() if code else "",
        }
        records.append(record)

    return records, skipped


def main():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    # Theory sheet
    print("\nProcessing Theory sheet...")
    theory_ws = wb['Theory']
    theory_records, t_skipped = process_theory(theory_ws)
    print(f"Processed {len(theory_records)} theory records (skipped {t_skipped})")

    with open(THEORY_JSON, 'w', encoding='utf-8') as f:
        json.dump(theory_records, f, indent=2, ensure_ascii=False)
    print(f"Saved Theory JSON to {THEORY_JSON}")

    # Practical sheet - use "Practical " (with trailing space, the cleaner student-facing one)
    # Try the cleaner sheet first, fallback to "Practical"
    practical_sheet_name = None
    for name in wb.sheetnames:
        if name.strip() == 'Practical':
            if practical_sheet_name is None:
                practical_sheet_name = name
            elif len(name) > len(practical_sheet_name):
                # Prefer the one with trailing space (cleaner format)
                pass
            else:
                practical_sheet_name = name

    # Use "Practical " (trailing space) which has cleaner student-facing columns
    for name in wb.sheetnames:
        if name == 'Practical ':
            practical_sheet_name = name
            break

    print(f"\nProcessing Practical sheet: '{practical_sheet_name}'...")
    practical_ws = wb[practical_sheet_name]
    practical_records, p_skipped = process_practical(practical_ws)
    print(f"Processed {len(practical_records)} practical records (skipped {p_skipped})")

    with open(PRACTICAL_JSON, 'w', encoding='utf-8') as f:
        json.dump(practical_records, f, indent=2, ensure_ascii=False)
    print(f"Saved Practical JSON to {PRACTICAL_JSON}")

    print(f"\n=== DONE ===")
    print(f"Theory:    {len(theory_records)} records")
    print(f"Practical: {len(practical_records)} records")
    print(f"Total:     {len(theory_records) + len(practical_records)} records")


if __name__ == "__main__":
    main()
