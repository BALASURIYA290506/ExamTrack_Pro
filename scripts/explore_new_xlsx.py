import openpyxl
import sys

XLSX_PATH = r'D:\DEV\Projects\ExamTrack_Pro\End Semester Theory & Practical Schedule(2025-26 EVEN & EVEN JUNIOR - Term2) - To be shared with the students.xlsx'

def main():
    try:
        wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
        print("Sheet names:", wb.sheetnames)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 3: # print first 3 rows to see headers
                    print([str(c) for c in row])
                else:
                    break
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    main()
